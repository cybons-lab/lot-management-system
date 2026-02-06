"""File handlers for master import.

Supports JSON, YAML, and Excel file formats.
CSV is explicitly NOT supported.
"""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING

import yaml

from app.presentation.schemas.import_schema import MasterImportRequest


if TYPE_CHECKING:
    from fastapi import UploadFile


class UnsupportedFileFormatError(Exception):
    """Raised when an unsupported file format is uploaded."""

    pass


class FileParseError(Exception):
    """Raised when file parsing fails."""

    pass


def get_file_extension(filename: str) -> str:
    """Extract file extension from filename."""
    return Path(filename).suffix.lower()


def parse_json(content: bytes) -> dict:
    """Parse JSON content."""
    try:
        parsed = json.loads(content.decode("utf-8"))
        if not isinstance(parsed, dict):
            raise FileParseError("JSON root must be an object")
        return parsed
    except json.JSONDecodeError as e:
        raise FileParseError(f"Invalid JSON: {e}")
    except UnicodeDecodeError as e:
        raise FileParseError(f"Invalid encoding (expected UTF-8): {e}")


def parse_yaml(content: bytes) -> dict:
    """Parse YAML content."""
    try:
        parsed = yaml.safe_load(content.decode("utf-8"))
        if not isinstance(parsed, dict):
            raise FileParseError("YAML root must be a mapping")
        return parsed
    except yaml.YAMLError as e:
        raise FileParseError(f"Invalid YAML: {e}")
    except UnicodeDecodeError as e:
        raise FileParseError(f"Invalid encoding (expected UTF-8): {e}")


def parse_excel(content: bytes) -> dict:
    """Parse Excel (.xlsx) content.

    Expects sheets named:
    - 'suppliers' for supply-side data
    - 'customers' for customer-side data

    Each sheet should have headers in the first row.
    """
    try:
        import openpyxl
    except ImportError:
        raise FileParseError(
            "openpyxl is required for Excel parsing. Install with: pip install openpyxl"
        )

    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise FileParseError(f"Failed to open Excel file: {e}")

    result: dict = {}

    # Parse suppliers sheet
    if "suppliers" in wb.sheetnames:
        result["supply_data"] = {"suppliers": _parse_suppliers_sheet(wb["suppliers"])}

    # Parse customers sheet
    if "customers" in wb.sheetnames:
        result["customer_data"] = {"customers": _parse_customers_sheet(wb["customers"])}

    if not result:
        template_sheet = None
        if "Template" in wb.sheetnames:
            template_sheet = wb["Template"]
        elif "Sheet1" in wb.sheetnames:
            template_sheet = wb["Sheet1"]

        if template_sheet:
            headers = [
                str(h).lower().strip() if h is not None else ""
                for h in next(template_sheet.iter_rows(values_only=True), [])
            ]
            # Check customer_code FIRST since customer templates may contain both
            # customer_code and supplier_code columns (supplier_code is used for items)
            if "customer_code" in headers:
                result["customer_data"] = {"customers": _parse_customers_sheet(template_sheet)}
            elif "supplier_code" in headers:
                result["supply_data"] = {"suppliers": _parse_suppliers_sheet(template_sheet)}

    wb.close()
    return result


def _parse_suppliers_sheet(sheet) -> list[dict]:
    """Parse suppliers sheet.

    Expected columns:
    - supplier_code, supplier_name, maker_part_code, product_name, is_primary, lead_time_days
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).lower() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    suppliers_dict: dict[str, dict] = {}

    for row in data_rows:
        if not row or all(cell is None for cell in row):
            continue

        row_dict = dict(zip(headers, row, strict=False))
        supplier_code = str(row_dict.get("supplier_code", "")).strip()
        if not supplier_code:
            continue

        if supplier_code not in suppliers_dict:
            suppliers_dict[supplier_code] = {
                "supplier_code": supplier_code,
                "supplier_name": str(row_dict.get("supplier_name", "")).strip(),
                "products": [],
            }

        # Add product if present
        maker_part_code = str(row_dict.get("maker_part_code", "")).strip()
        if maker_part_code:
            suppliers_dict[supplier_code]["products"].append(
                {
                    "maker_part_code": maker_part_code,
                    "product_name": str(row_dict.get("product_name", "")).strip() or None,
                    "is_primary": _to_bool(row_dict.get("is_primary", False)),
                    "lead_time_days": _to_int(row_dict.get("lead_time_days")),
                }
            )

    return list(suppliers_dict.values())


def _parse_customers_sheet(sheet) -> list[dict]:
    """Parse customers sheet.

    Expected columns:
    - customer_code, customer_name, delivery_place_code, delivery_place_name,
    - external_product_code, maker_part_code, supplier_code
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).lower() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    customers_dict: dict[str, dict] = {}

    for row in data_rows:
        if not row or all(cell is None for cell in row):
            continue

        row_dict = dict(zip(headers, row, strict=False))
        customer_code = str(row_dict.get("customer_code", "")).strip()
        if not customer_code:
            continue

        if customer_code not in customers_dict:
            customers_dict[customer_code] = {
                "customer_code": customer_code,
                "customer_name": str(row_dict.get("customer_name", "")).strip(),
                "delivery_places": [],
                "items": [],
            }

        customer = customers_dict[customer_code]

        # Add delivery place if present
        dp_code = str(row_dict.get("delivery_place_code", "")).strip()
        if dp_code and not any(
            dp["delivery_place_code"] == dp_code for dp in customer["delivery_places"]
        ):
            customer["delivery_places"].append(
                {
                    "delivery_place_code": dp_code,
                    "delivery_place_name": str(row_dict.get("delivery_place_name", "")).strip(),
                    "jiku_code": str(row_dict.get("jiku_code", "")).strip() or None,
                }
            )

        # Add customer item if present
        ext_code = (
            str(row_dict.get("customer_part_no", "")).strip()
            or str(row_dict.get("external_product_code", "")).strip()
        )
        if ext_code:
            customer["items"].append(
                {
                    "customer_part_no": ext_code,
                    "maker_part_code": str(row_dict.get("maker_part_code", "")).strip(),
                    "supplier_code": str(row_dict.get("supplier_code", "")).strip() or None,
                }
            )

    return list(customers_dict.values())


def _to_bool(value) -> bool:
    """Convert value to boolean."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ("true", "yes", "1", "on")
    return bool(value)


def _to_int(value) -> int | None:
    """Convert value to int or None."""
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


async def parse_import_file(file: UploadFile) -> MasterImportRequest:
    """Parse uploaded file and return MasterImportRequest.

    Supports: .json, .yaml, .yml, .xlsx
    Does NOT support: .csv (will raise UnsupportedFileFormatError)
    """
    if not file.filename:
        raise FileParseError("Filename is required")

    ext = get_file_extension(file.filename)

    # Explicitly reject CSV
    if ext == ".csv":
        raise UnsupportedFileFormatError(
            "CSV format is not supported. Please use Excel (.xlsx), JSON, or YAML."
        )

    content = await file.read()

    if ext == ".json":
        data = parse_json(content)
    elif ext in (".yaml", ".yml"):
        data = parse_yaml(content)
    elif ext == ".xlsx":
        data = parse_excel(content)
    else:
        raise UnsupportedFileFormatError(
            f"Unsupported file format: {ext}. Supported formats: .xlsx, .json, .yaml, .yml"
        )

    return MasterImportRequest.model_validate(data)
