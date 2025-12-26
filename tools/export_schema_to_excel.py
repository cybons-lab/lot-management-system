# tools/export_schema_to_excel.py
# Python 3.13 / 依存: pandas, SQLAlchemy, openpyxl
import argparse
import re
from pathlib import Path
from typing import Dict, List

import pandas as pd
from sqlalchemy import create_engine, text


# ===== デフォルト（引数なしで動く） =====
DEFAULT_URL = "postgresql://admin:dev_password@localhost:5432/lot_management"
DEFAULT_SCHEMA = "public"
DEFAULT_OUT = "docs/schema/schema_definition.xlsx"

# Excelテーブルの薄いグリーン（UI最下段の薄色）
TABLE_STYLE_NAME = "TableStyleLight21"

# タブ色（ARP：RGB）
TABCOLOR_CONSTRAINTS = "FFC000"  # 橙
TABCOLOR_FOREIGN_KEYS = "00B0F0"  # 青


def sanitize_sheet_name(name: str) -> str:
    # 禁止文字 \ / * ? : [ ] を除去
    name = re.sub(r"[\\/*?:\[\]]", "_", name)
    # 31文字に切り捨て
    return name[:31] if len(name) > 31 else name


def unique_sheet_name(name: str, used: set) -> str:
    base = sanitize_sheet_name(name)
    if base not in used:
        used.add(base)
        return base
    # 重複した場合に _1, _2 ... を付与（31文字制限考慮）
    i = 1
    while True:
        suffix = f"_{i}"
        trimmed = base[: max(0, 31 - len(suffix))] + suffix
        if trimmed not in used:
            used.add(trimmed)
            return trimmed
        i += 1


def get_tables(conn, schema: str) -> List[str]:
    sql = text("""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = :schema AND table_type = 'BASE TABLE'
        ORDER BY table_name
    """)
    return [r[0] for r in conn.execute(sql, {"schema": schema}).fetchall()]


def get_table_comment(conn, schema: str, table: str) -> str:
    sql = text("""
        SELECT obj_description(cls.oid) AS comment
        FROM pg_class cls
        JOIN pg_namespace nsp ON nsp.oid = cls.relnamespace
        WHERE nsp.nspname = :schema AND cls.relname = :table
        LIMIT 1
    """)
    row = conn.execute(sql, {"schema": schema, "table": table}).fetchone()
    return row[0] if row and row[0] else ""


def get_columns_df(conn, schema: str, table: str) -> pd.DataFrame:
    sql = text("""
    WITH pk_cols AS (
        SELECT kcu.column_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu
          ON tc.constraint_name = kcu.constraint_name
         AND tc.table_schema = kcu.table_schema
         AND tc.table_name   = kcu.table_name
        WHERE tc.table_schema = :schema
          AND tc.table_name   = :table
          AND tc.constraint_type = 'PRIMARY KEY'
    ),
    fk_cols AS (
        SELECT DISTINCT kcu.column_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu
          ON tc.constraint_name = kcu.constraint_name
         AND tc.table_schema = kcu.table_schema
         AND tc.table_name   = kcu.table_name
        WHERE tc.table_schema = :schema
          AND tc.table_name   = :table
          AND tc.constraint_type = 'FOREIGN KEY'
    )
    SELECT
        cols.ordinal_position   AS "No",
        cols.column_name        AS "列名",
        cols.data_type          AS "データ型",
        COALESCE(cols.character_maximum_length::text,
                 CASE
                   WHEN cols.numeric_precision IS NOT NULL AND cols.numeric_scale IS NOT NULL
                     THEN cols.numeric_precision::text || ',' || cols.numeric_scale::text
                   WHEN cols.numeric_precision IS NOT NULL
                     THEN cols.numeric_precision::text
                 END, '')        AS "長さ/精度",
        CASE WHEN cols.is_nullable = 'NO' THEN 'NOT NULL' ELSE '' END AS "NULL制",
        COALESCE(cols.column_default, '') AS "デフォルト",
        CASE WHEN cols.column_name IN (SELECT column_name FROM pk_cols) THEN 'PK' ELSE '' END AS "PK",
        CASE WHEN cols.column_name IN (SELECT column_name FROM fk_cols) THEN 'FK' ELSE '' END AS "FK",
        COALESCE(pg_catalog.col_description(cls.oid, cols.ordinal_position), '') AS "列コメント"
    FROM information_schema.columns cols
    JOIN pg_class cls
      ON cls.relname = cols.table_name
    JOIN pg_namespace nsp
      ON nsp.oid = cls.relnamespace AND nsp.nspname = cols.table_schema
    WHERE cols.table_schema = :schema AND cols.table_name = :table
    ORDER BY cols.ordinal_position
    """)
    return pd.read_sql(sql, conn, params={"schema": schema, "table": table})


def get_constraints_df(conn, schema: str) -> pd.DataFrame:
    sql = text("""
    SELECT
      n.nspname          AS schema,
      c.relname          AS table_name,
      con.conname        AS constraint_name,
      CASE con.contype
        WHEN 'p' THEN 'PRIMARY KEY'
        WHEN 'u' THEN 'UNIQUE'
        WHEN 'f' THEN 'FOREIGN KEY'
        WHEN 'c' THEN 'CHECK'
        ELSE con.contype::text
      END                AS constraint_type,
      pg_get_constraintdef(con.oid, true) AS definition
    FROM pg_constraint con
    JOIN pg_class c      ON c.oid = con.conrelid
    JOIN pg_namespace n  ON n.oid = c.relnamespace
    WHERE n.nspname = :schema
    ORDER BY table_name, constraint_type, constraint_name
    """)
    df = pd.read_sql(sql, conn, params={"schema": schema})
    return df.rename(
        columns={
            "schema": "スキーマ",
            "table_name": "テーブル名",
            "constraint_name": "制約名",
            "constraint_type": "制約種別",
            "definition": "定義",
        }
    )[["スキーマ", "テーブル名", "制約名", "制約種別", "定義"]]


def get_foreign_keys_df(conn, schema: str) -> pd.DataFrame:
    sql = text("""
    SELECT
      rc.constraint_name                          AS fk_name,
      kcu.table_name                               AS src_table,
      kcu.column_name                              AS src_column,
      ccu.table_name                               AS dst_table,
      ccu.column_name                              AS dst_column,
      rc.update_rule,
      rc.delete_rule
    FROM information_schema.referential_constraints rc
    JOIN information_schema.key_column_usage kcu
      ON kcu.constraint_name = rc.constraint_name
     AND kcu.constraint_schema = rc.constraint_schema
    JOIN information_schema.constraint_column_usage ccu
      ON ccu.constraint_name = rc.unique_constraint_name
     AND ccu.constraint_schema = rc.unique_constraint_schema
    WHERE rc.constraint_schema = :schema
    ORDER BY src_table, fk_name, kcu.ordinal_position
    """)
    df = pd.read_sql(sql, conn, params={"schema": schema})
    return df.rename(
        columns={
            "fk_name": "FK名",
            "src_table": "参照元テーブル",
            "src_column": "参照元列",
            "dst_table": "参照先テーブル",
            "dst_column": "参照先列",
            "update_rule": "更新時",
            "delete_rule": "削除時",
        }
    )


def add_table_style(ws, df: pd.DataFrame, table_name: str, start_row=1, start_col=1):
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font

    end_row = start_row + len(df.index)
    end_col = start_col + len(df.columns) - 1
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

    tbl = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name=TABLE_STYLE_NAME,  # 最下段の薄いグリーン
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    # 列幅およびヘッダー可読性
    for i, col in enumerate(df.columns, start=start_col):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col] if v is not None])
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, max_len + 2), 60)

    # ヘッダー行は太字・折返し・中央寄せ
    header_row = ws[
        f"{get_column_letter(start_col)}{start_row}" : f"{get_column_letter(end_col)}{start_row}"
    ][0]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = ws["A2"]  # 1行目固定


def build_index_sheet(xw, sheet_order: List[str]):
    """最初のシートに目次（ハイパーリンク）を作成"""
    ws = xw.book.create_sheet(title="INDEX", index=0)
    ws["A1"] = "テーブル一覧"
    ws["A2"] = "クリックで各シートへ移動します。"

    row = 4
    for name in sheet_order:
        # =HYPERLINK("#'name'!A1","name")
        ws.cell(row=row, column=1).value = f'=HYPERLINK("#\'{name}\'!A1","{name}")'
        row += 1

    ws.freeze_panes = ws["A4"]
    ws.column_dimensions["A"].width = 50


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", default=DEFAULT_URL, help="SQLAlchemy URL")
    ap.add_argument("--schema", default=DEFAULT_SCHEMA)
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    engine = create_engine(args.url, future=True)

    with engine.connect() as conn, pd.ExcelWriter(args.out, engine="openpyxl") as xw:
        tables = get_tables(conn, args.schema)

        used_names = set()
        sheet_map: Dict[str, str] = {}  # original_table_name -> sheet_name
        ordered_sheet_names: List[str] = []

        # テーブル別シート
        for tbl in tables:
            df = get_columns_df(conn, args.schema, tbl)

            # シート名生成（31文字制限＆重複回避）
            sheet_name = unique_sheet_name(tbl, used_names)
            sheet_map[tbl] = sheet_name
            ordered_sheet_names.append(sheet_name)

            # メタ行（テーブル名・コメント）
            meta_df = pd.DataFrame(
                [
                    {
                        "テーブル名": tbl,
                        "テーブルコメント": get_table_comment(conn, args.schema, tbl),
                    }
                ]
            )

            meta_df.to_excel(xw, sheet_name=sheet_name, index=False, startrow=0)
            df.to_excel(xw, sheet_name=sheet_name, index=False, startrow=2)

            ws = xw.book[sheet_name]
            # Excelテーブル化（薄グリーン）
            add_table_style(
                ws,
                df,
                table_name=f"T_{re.sub('[^A-Za-z0-9_]', '_', tbl)}",
                start_row=3,
                start_col=1,
            )

        # 制約一覧
        constraints_df = get_constraints_df(conn, args.schema)
        constraints_df.to_excel(xw, sheet_name="CONSTRAINTS", index=False)
        add_table_style(
            xw.book["CONSTRAINTS"],
            constraints_df,
            table_name="T_CONSTRAINTS",
            start_row=1,
            start_col=1,
        )
        xw.book["CONSTRAINTS"].sheet_properties.tabColor = TABCOLOR_CONSTRAINTS

        # 外部キー一覧
        fks_df = get_foreign_keys_df(conn, args.schema)
        fks_df.to_excel(xw, sheet_name="FOREIGN_KEYS", index=False)
        add_table_style(
            xw.book["FOREIGN_KEYS"],
            fks_df,
            table_name="T_FOREIGN_KEYS",
            start_row=1,
            start_col=1,
        )
        xw.book["FOREIGN_KEYS"].sheet_properties.tabColor = TABCOLOR_FOREIGN_KEYS

        # 目次（先頭）
        build_index_sheet(xw, ordered_sheet_names + ["CONSTRAINTS", "FOREIGN_KEYS"])

    print(f"[OK] Exported: {args.out}")


if __name__ == "__main__":
    main()
